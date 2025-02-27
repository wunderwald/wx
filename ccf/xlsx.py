from openpyxl import Workbook, load_workbook

def write_xlsx(vectors: dict, single_values: dict, output_path: str):
    """
    Write data to an Excel file.
    Args:
        vectors (dict): A dictionary where keys are column names and values are lists of time series data.
        single_value (dict): A dictionary where keys are column names and values are single data values.
        output_path (str): The file path where the Excel file will be saved.
    Returns:
        None
    """

    # Create a new workbook
    wb = Workbook()
    sheet = wb.active
    sheet.title = "(w)ccf data"

    # Write the single valued data to columns 
    # (column 1 -> names, column 2 -> values)
    row_index = 1
    for name, value in single_values.items():
        sheet.cell(row=row_index, column=1, value=name)  
        sheet.cell(row=row_index, column=2, value=str(value))  
        row_index += 1
    
    # Write the time series data to columns
    # (one vector with title per column, starting in column 4)
    column_index = 4
    for name in vectors.keys():
        sheet.cell(row=1, column=column_index, value=name)
        for row_index, value in enumerate(vectors[name]):  
            sheet.cell(row=row_index+2, column=column_index, value=str(value))
        column_index += 1

    # Save the workbook
    wb.save(output_path)

def read_xlsx(input_path: str):
    """
    Read data from xlsx file and create openpyxl workbook object.
    Args:
        input_path (str): The file path where the Excel file is saved.
    Returns:
        wb (Workbook): openpyxl workbook object
    """
    wb = load_workbook(input_path)
    return wb

def get_sheet_names(workbook: Workbook):
    """
    Get the names of the sheets in the workbook.
    Args:
        workbook (Workbook): openpyxl workbook object
    Returns:
        sheet_names (list): List of sheet names
    """
    sheet_names = workbook.sheetnames
    return sheet_names

def get_columns(workbook: Workbook, sheet_name: str, headers: bool=True):
    """
    Get the columns from the sheet in the workbook.
    Args:
        workbook (Workbook): openpyxl workbook object
        sheet_name (str): The name of the sheet
        headers (bool): If True, the first row is considered as headers
    Returns:
        columns (dict): A dictionary where keys are column names and values are lists of data
    """
    sheet = workbook[sheet_name]
    columns = {}
    for index, column in enumerate(sheet.iter_cols(values_only=True)):
        if headers:
            column_name = f"{column[0]}"
            column_data = column[1:]
        else:
            column_name = f"Column_{index+1}"
            column_data = column
        columns[column_name] = column_data
    return columns
    