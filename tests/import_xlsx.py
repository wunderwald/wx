import openpyxl

workbook = openpyxl.load_workbook('data/HRV_1009_127_FS_C.xlsx')

# select sheet
sheet_names = workbook.sheetnames
print("Sheet Names:", sheet_names)

sheet = workbook.active
print('Active Sheet:', sheet.title)

# Read data from the sheet
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)

# Print the data
for row in data:
    print(row)

